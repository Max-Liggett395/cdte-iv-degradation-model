#!/usr/bin/env python3
"""
build_master_workbook.py  --  src/io/

Build ONE Excel workbook holding everything parsed from the raw Solmetric
exports: per-module/per-timepoint metadata, measured + model-predicted
electricals, both environment flavors, derived quantities, and STC-corrected
values. Corrections are written as live Excel formulas that reference an
Assumptions sheet, so changing a temperature coefficient recomputes the book.

Run from repo root:
    python src/io/build_master_workbook.py            # uses data/raw_iv_traces
    python src/io/build_master_workbook.py <data_dir> <out.xlsx>
"""
import os
import sys
import datetime as dt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parse_iv_master import parse_directory  # noqa: E402

# ---- datasheet temperature coefficients (relative, per-degree-C) ----
# Source: First Solar Series 7 TR1 Module Datasheet (FS-7XXXA-TR1).
#   TK(Pmax) = -0.32 %/C ;  TK(Voc) = -0.28 %/C ;  TK(Isc) = +0.04 %/C
COEFFS = {"G_ref": 1000.0, "T_ref": 25.0,
          "TK_Pmax": -0.0032, "TK_Voc": -0.0028, "TK_Isc": 0.0004}

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
GRP_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FONT = Font(name=ARIAL, color="0000FF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
BASE_FONT = Font(name=ARIAL, size=10)
TITLE_FONT = Font(name=ARIAL, bold=True, size=14)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Master column spec: (header, parser_key|None, kind, number_format)
# kind: 'val' = value from parser, 'fml' = formula template (use {r})
MASTER_COLS = [
    ("meas_id", "meas_id", "val", "0"),
    ("module", "module", "val", None),
    ("bay", "bay", "val", "0"),
    ("module_id", "module_id", "val", None),
    ("config", "config", "val", None),
    ("quality", "quality", "val", None),
    ("datetime_local", "datetime_local", "val", "yyyy-mm-dd hh:mm"),
    ("tz_offset_hr", "tz_offset_hr", "val", "0"),
    ("note", "note", "val", None),
    ("module_model", "module_model", "val", None),
    ("array_location", "array_location", "val", None),
    ("project_file", "project_file", "val", None),
    ("irr_solsensor_wm2", "irr_solsensor_wm2", "val", "0.0"),
    ("temp_tc1_c", "temp_tc1_c", "val", "0.00"),
    ("temp_tc2_c", "temp_tc2_c", "val", "0.00"),
    ("tilt_meas_deg", "tilt_meas_deg", "val", "0.0"),
    ("irr_model_wm2", "irr_model_wm2", "val", "0.0"),
    ("cell_temp_model_c", "cell_temp_model_c", "val", "0.00"),
    ("aoi_deg", "aoi_deg", "val", "0.0"),
    ("array_azimuth_deg", "array_azimuth_deg", "val", "0"),
    ("user_series_r_ohm", "user_series_r_ohm", "val", "0.0000"),
    ("pmax_meas_w", "pmax_meas_w", "val", "0.00"),
    ("vmpp_meas_v", "vmpp_meas_v", "val", "0.00"),
    ("impp_meas_a", "impp_meas_a", "val", "0.0000"),
    ("voc_meas_v", "voc_meas_v", "val", "0.00"),
    ("isc_meas_a", "isc_meas_a", "val", "0.0000"),
    ("pmax_model_w", "pmax_model_w", "val", "0.00"),
    ("vmpp_model_v", "vmpp_model_v", "val", "0.00"),
    ("impp_model_a", "impp_model_a", "val", "0.0000"),
    ("voc_model_v", "voc_model_v", "val", "0.00"),
    ("isc_model_a", "isc_model_a", "val", "0.0000"),
    ("performance_factor_pct", "performance_factor_pct", "val", "0.0"),
    ("n_iv_points", "n_iv_points", "val", "0"),
    # derived
    ("ff_meas", None, "=V{r}/(Y{r}*Z{r})", "0.0000"),
    ("irr_derate_ratio", None, "=Q{r}/M{r}", "0.0000"),
    ("smarttemp_offset_c", None, "=R{r}-N{r}", "0.00"),
    # STC-corrected (effective irradiance Q, cell temp R, Assumptions coeffs)
    ("isc_stc_a", None,
     "=Z{r}*(Assumptions!$B$2/Q{r})/(1+Assumptions!$B$6*(R{r}-Assumptions!$B$3))",
     "0.0000"),
    ("voc_stc_v", None,
     "=Y{r}/(1+Assumptions!$B$5*(R{r}-Assumptions!$B$3))", "0.00"),
    ("pmax_stc_w", None,
     "=V{r}*(Assumptions!$B$2/Q{r})/(1+Assumptions!$B$4*(R{r}-Assumptions!$B$3))",
     "0.00"),
    ("ff_stc", None, "=AM{r}/(AL{r}*AK{r})", "0.0000"),
]

DICT_ROWS = [
    ("meas_id", "", "Unique row id per measurement; joins to IV_Curves"),
    ("module", "", "Module label: B<bay> <module_id>[-OC/SC]"),
    ("config", "", "OC / SC = open- vs closed-circuit module (Bay 1); separate physical modules"),
    ("quality", "", "Location tag parsed from Array Location (bad / bad fav / north end ...)"),
    ("datetime_local", "local clock", "Measurement timestamp in the file's local time zone"),
    ("tz_offset_hr", "h", "GMT offset of the local timestamp"),
    ("irr_solsensor_wm2", "W/m^2", "Raw plane-of-array irradiance from SolSensor"),
    ("temp_tc1_c", "deg C", "Back-of-module thermocouple temperature"),
    ("irr_model_wm2", "W/m^2", "Effective irradiance the PVA model used (AOI/reflection-derated)"),
    ("cell_temp_model_c", "deg C", "SmartTemp estimated cell temperature used by the PVA model"),
    ("aoi_deg", "deg", "Angle of incidence at measurement"),
    ("user_series_r_ohm", "ohm", "User-entered series resistance (lead/wire)"),
    ("pmax_meas_w/.._model_w", "W", "Measured vs native-software model-predicted maximum power"),
    ("voc_meas_v/.._model_v", "V", "Measured vs model-predicted open-circuit voltage"),
    ("isc_meas_a/.._model_a", "A", "Measured vs model-predicted short-circuit current"),
    ("performance_factor_pct", "%", "PVA software: 100 x measured Pmax / model-predicted Pmax"),
    ("ff_meas", "", "Fill factor = Pmax_meas / (Voc_meas x Isc_meas)  [formula]"),
    ("irr_derate_ratio", "", "irr_model / irr_solsensor  [formula]"),
    ("smarttemp_offset_c", "deg C", "cell_temp_model - temp_tc1  [formula]"),
    ("isc_stc_a", "A", "Isc corrected to STC using effective irradiance + cell temp + TK_Isc  [formula]"),
    ("voc_stc_v", "V", "Voc corrected to 25 C using cell temp + TK_Voc  [formula]"),
    ("pmax_stc_w", "W", "Pmax corrected to STC: irradiance scale + TK_Pmax temperature term  [formula]"),
    ("ff_stc", "", "Fill factor of the STC-corrected operating point  [formula]"),
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def build(data_dir, out_path):
    master, iv = parse_directory(data_dir, recursive=True)
    n = len(master)
    last = n + 1  # last data row in Master (header = row 1)

    wb = Workbook()

    # ---------------- README ----------------
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "CdTe Field IV — Master Workbook"
    ws["A1"].font = TITLE_FONT
    notes = [
        "",
        f"Generated:  {dt.datetime.now():%Y-%m-%d %H:%M}",
        f"Source dir: {data_dir}",
        f"Measurements parsed: {n}     IV points: {len(iv)}",
        "",
        "SHEETS",
        "  Master         One row per module per timepoint: metadata, measured +",
        "                 model-predicted electricals, both irradiance/temperature",
        "                 flavors, derived quantities, and STC-corrected values.",
        "  Module_Summary Per-module rollups (count, date span, mean/min/max).",
        "  IV_Curves      Full raw IV traces (V, I, P), joinable via meas_id.",
        "  Assumptions    Reference conditions + datasheet temperature coefficients.",
        "  Data_Dictionary Column definitions and units.",
        "",
        "CORRECTIONS",
        "  STC columns use the MODEL-EFFECTIVE irradiance (irr_model_wm2) and the",
        "  SmartTemp cell temperature (cell_temp_model_c) — the same inputs the PVA",
        "  software uses — not the raw SolSensor / thermocouple values. They are live",
        "  Excel formulas referencing the Assumptions sheet; edit a coefficient there",
        "  and the whole workbook recomputes.",
        "",
        "  Coefficients are RELATIVE (%/C). Source: First Solar Series 7 TR1 datasheet.",
        "  Confirm against the FS-7520A bin once the datasheet PDF is in hand.",
        "",
        "REPRODUCE",
        "  python src/io/build_master_workbook.py <data_dir> <out.xlsx>",
    ]
    for i, t in enumerate(notes, start=2):
        ws[f"A{i}"] = t
        ws[f"A{i}"].font = BASE_FONT
    ws.column_dimensions["A"].width = 95

    # ---------------- Assumptions ----------------
    a = wb.create_sheet("Assumptions")
    a.sheet_view.showGridLines = False
    a["A1"] = "Parameter"; a["B1"] = "Value"; a["C1"] = "Units"; a["D1"] = "Source"
    style_header(a, 1, 4)
    arows = [
        ("G_ref (reference irradiance)", COEFFS["G_ref"], "W/m^2", "STC"),
        ("T_ref (reference temperature)", COEFFS["T_ref"], "deg C", "STC"),
        ("TK_Pmax", COEFFS["TK_Pmax"], "1/deg C", "First Solar Series 7 TR1 datasheet (-0.32 %/C)"),
        ("TK_Voc", COEFFS["TK_Voc"], "1/deg C", "First Solar Series 7 TR1 datasheet (-0.28 %/C)"),
        ("TK_Isc", COEFFS["TK_Isc"], "1/deg C", "First Solar Series 7 TR1 datasheet (+0.04 %/C)"),
    ]
    for i, (p, v, u, s) in enumerate(arows, start=2):
        a.cell(i, 1, p).font = BASE_FONT
        cv = a.cell(i, 2, v); cv.font = INPUT_FONT; cv.fill = INPUT_FILL
        cv.number_format = "0.0000"
        a.cell(i, 3, u).font = BASE_FONT
        a.cell(i, 4, s).font = BASE_FONT
    for col, w in zip("ABCD", (30, 12, 10, 52)):
        a.column_dimensions[col].width = w

    # ---------------- Master ----------------
    m = wb.create_sheet("Master")
    headers = [c[0] for c in MASTER_COLS]
    for j, h in enumerate(headers, start=1):
        m.cell(1, j, h)
    style_header(m, 1, len(headers))
    m.freeze_panes = "C2"

    for ridx in range(n):
        r = ridx + 2
        rec = master.iloc[ridx]
        for j, (hdr, key, kind, fmt) in enumerate(MASTER_COLS, start=1):
            cell = m.cell(row=r, column=j)
            if kind == "val":
                val = rec[key]
                if isinstance(val, float) and np.isnan(val):
                    val = None
                if isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
                cell.value = val
            else:  # formula
                cell.value = kind.format(r=r)
            cell.font = BASE_FONT
            if fmt:
                cell.number_format = fmt
    # widths
    for j, (hdr, *_rest) in enumerate(MASTER_COLS, start=1):
        L = get_column_letter(j)
        m.column_dimensions[L].width = max(9, min(20, len(hdr) + 2))

    # ---------------- Module_Summary ----------------
    s = wb.create_sheet("Module_Summary")
    s.sheet_view.showGridLines = False
    scols = ["module", "n", "first_date", "last_date",
             "mean_PF_pct", "min_PF_pct", "max_PF_pct",
             "mean_ff_meas", "mean_pmax_meas_w", "mean_pmax_stc_w", "mean_ff_stc"]
    for j, h in enumerate(scols, start=1):
        s.cell(1, j, h)
    style_header(s, 1, len(scols))
    s.freeze_panes = "B2"
    mods = list(dict.fromkeys(master["module"].tolist()))
    Bm = f"Master!$B$2:$B${last}"
    G = f"Master!$G$2:$G${last}"
    AF = f"Master!$AF$2:$AF${last}"
    AH = f"Master!$AH$2:$AH${last}"
    V = f"Master!$V$2:$V${last}"
    AM = f"Master!$AM$2:$AM${last}"
    AN = f"Master!$AN$2:$AN${last}"
    for i, mod in enumerate(mods, start=2):
        k = f"$A{i}"
        s.cell(i, 1, mod).font = BASE_FONT
        s.cell(i, 2, f'=COUNTIF({Bm},{k})')
        s.cell(i, 3, f'=_xlfn.MINIFS({G},{Bm},{k})').number_format = "yyyy-mm-dd"
        s.cell(i, 4, f'=_xlfn.MAXIFS({G},{Bm},{k})').number_format = "yyyy-mm-dd"
        s.cell(i, 5, f'=AVERAGEIFS({AF},{Bm},{k})').number_format = "0.0"
        s.cell(i, 6, f'=_xlfn.MINIFS({AF},{Bm},{k})').number_format = "0.0"
        s.cell(i, 7, f'=_xlfn.MAXIFS({AF},{Bm},{k})').number_format = "0.0"
        s.cell(i, 8, f'=AVERAGEIFS({AH},{Bm},{k})').number_format = "0.0000"
        s.cell(i, 9, f'=AVERAGEIFS({V},{Bm},{k})').number_format = "0.0"
        s.cell(i, 10, f'=AVERAGEIFS({AM},{Bm},{k})').number_format = "0.0"
        s.cell(i, 11, f'=AVERAGEIFS({AN},{Bm},{k})').number_format = "0.0000"
        for j in range(2, len(scols) + 1):
            s.cell(i, j).font = BASE_FONT
    for col, w in zip("ABCDEFGHIJK",
                      (14, 6, 12, 12, 11, 10, 10, 12, 16, 15, 11)):
        s.column_dimensions[col].width = w

    # ---------------- IV_Curves ----------------
    c = wb.create_sheet("IV_Curves")
    for j, h in enumerate(["meas_id", "module", "v", "i", "p"], start=1):
        c.cell(1, j, h)
    style_header(c, 1, 5)
    c.freeze_panes = "A2"
    if len(iv):
        for rec in iv.itertuples(index=False):
            c.append([rec.meas_id, rec.module, rec.v, rec.i, rec.p])
    for col, w in zip("ABCDE", (9, 14, 14, 12, 14)):
        c.column_dimensions[col].width = w

    # ---------------- Data_Dictionary ----------------
    d = wb.create_sheet("Data_Dictionary")
    d.sheet_view.showGridLines = False
    for j, h in enumerate(["column", "units", "description"], start=1):
        d.cell(1, j, h)
    style_header(d, 1, 3)
    for i, (col, units, desc) in enumerate(DICT_ROWS, start=2):
        d.cell(i, 1, col).font = BASE_FONT
        d.cell(i, 2, units).font = BASE_FONT
        d.cell(i, 3, desc).font = BASE_FONT
    for col, w in zip("ABC", (26, 12, 90)):
        d.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path, n, len(iv), mods


if __name__ == "__main__":
    data_dir = sys.argv[1] if len(sys.argv) > 1 else "data/raw_iv_traces"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "outputs/master_iv_workbook.xlsx"
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    p, n, nv, mods = build(data_dir, out_path)
    print(f"Wrote {p}: {n} measurements, {nv} IV points, {len(mods)} modules.")